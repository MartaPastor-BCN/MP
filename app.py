import streamlit as st
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors

# ================================================================================
# PAGE CONFIG & STYLING
# ================================================================================
st.set_page_config(
    page_title="Media Planner Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    .main { padding: 2rem 1.5rem; }
    .stButton > button { 
        padding: 0.6rem 1rem; 
        font-weight: 500; 
        border-radius: 6px;
        transition: all 0.3s ease;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
    }
    h2 { 
        color: #0078D4; 
        font-size: 1.4rem; 
        margin-top: 2.5rem; 
        margin-bottom: 1rem; 
        border-bottom: none; 
        padding-bottom: 0;
        letter-spacing: 0.3px;
        font-weight: 600;
    }
    h3 {
        letter-spacing: 0.2px;
        font-weight: 600;
    }
    .info-box { 
        background-color: #FAFAFA; 
        padding: 14px; 
        border-left: 3px solid #0078D4; 
        border-radius: 4px; 
        margin: 1.5rem 0; 
        color: #333; 
        font-size: 0.95rem;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
    }
    
    /* Table Styling - Subtle Shadows */
    .stDataFrame {
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.06) !important;
        border: none !important;
    }
    
    /* Spacing */
    .section-spacer {
        margin-top: 2rem;
        margin-bottom: 2rem;
    }
    
    /* Metric Cards */
    .metric-card {
        background: #FAFAFA;
        border-radius: 6px;
        padding: 0;
        border-left: 3px solid #0078D4;
    }
    
    /* Animations */
    @keyframes fadeIn {
        from {
            opacity: 0;
            transform: translateY(-5px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    @keyframes slideUp {
        from {
            opacity: 0;
            transform: translateY(10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .metric-value {
        animation: slideUp 0.5s ease-out;
    }
    
    /* Header Gradient */
    .section-header {
        background: linear-gradient(135deg, #0078D4 0%, #0063B1 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        letter-spacing: 0.5px;
    }
</style>
""", unsafe_allow_html=True)

# ================================================================================
# 1. HEADER / VERSIONING
# ================================================================================
st.markdown("# Media Planner Tool")
st.markdown("**Monetize Guaranteed - Seat 280** | Version 2.0")

st.markdown("""
<div class="info-box">
    <strong>Guaranteed delivery planning</strong> • CPM rates should align with DNV Rate Card • Seat 280 only
</div>
""", unsafe_allow_html=True)

# ================================================================================
# INITIALIZE SESSION STATE
# ================================================================================
if "flights" not in st.session_state:
    st.session_state.flights = []
if "campaign_details" not in st.session_state:
    st.session_state.campaign_details = {}

# ================================================================================
# PRODUCT RULES - MONETIZE GUARANTEED ONLY (SEAT 280)
# Full matrix: publisher × format × device — no PMPs
# ================================================================================
PRODUCT_RULES = {
    # ── PG FIRST IMPRESSION ──────────────────────────────────────────────────
    "PG - First Impression": {
        "seat": "280 - Monetize",
        "line_item_type": "PG",
        "revenue_type": "CPM",
        "delivery_type": "N/A",
        "allow_rtb": False,
        "underspend_catchup": "N/A",
        "impression_pacing": "100%",
        "creative_rotation": "N/A",
        "landing_page": "None",
        "publishers": ["MSN"],
        "formats": ["Banner"],
        "priority": 7,
        "frequency_cap": "1 imp/day",
        "pacing": "ASAP",
        "inventory_type": "Banner",
        "geo_targeting": "Country Targeting",
        "publisher_targeting": "KV pub = msn",
        "supply_targeting": "App & Web",
        "device_config": {
            "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet",
                            "ad_sizes": "300x250, 300x600, 728x90, 970x250, 320x50",
                            "inventory_targeting": "KV pub = msn", "min_creatives": 1},
            "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",
                            "ad_sizes": "728x90, 970x250, 300x250, 300x600",
                            "inventory_targeting": "KV pub = msn", "min_creatives": 1},
            "Mobile":      {"supply": "Web Only",  "device_targeting": "Mobile, Tablet",
                            "ad_sizes": "320x50",
                            "inventory_targeting": "KV pub = msn", "min_creatives": 1},
        },
    },
    # ── PG STANDARD ──────────────────────────────────────────────────────────
    "PG - Standard": {
        "seat": "280 - Monetize",
        "line_item_type": "PG",
        "revenue_type": "CPM",
        "delivery_type": "N/A",
        "allow_rtb": False,
        "underspend_catchup": "N/A",
        "impression_pacing": "100%",
        "creative_rotation": "N/A",
        "landing_page": "None",
        "publishers": ["MSN", "Outlook", "MCG"],
        "formats": ["Banner", "Native", "Video"],
        "priority": 5,
        "frequency_cap": "6 imps/day",
        "pacing": "Even",
        "inventory_type": "Banner",
        "geo_targeting": "Country Targeting",
        "publisher_targeting": "KV pub = msn / outlook / microsoftcasualgames",
        "supply_targeting": "App & Web",
        "publisher_format_config": {
            ("MSN", "Banner"): {
                "inventory_targeting": "KV pub = msn",
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "300x250, 300x600, 728x90, 970x250, 320x50"},
                    "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",          "ad_sizes": "728x90, 970x250, 300x250, 300x600"},
                    "Mobile":      {"supply": "Web Only",  "device_targeting": "Mobile, Tablet",           "ad_sizes": "320x50"},
                },
            },
            ("MSN", "Video"): {
                "inventory_targeting": "KV pub = msn",
                "creative_specs": "15 seconds OR 30 seconds",
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "N/A"},
                    "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",          "ad_sizes": "N/A"},
                    "Mobile":      {"supply": "Web Only",  "device_targeting": "Mobile, Tablet",           "ad_sizes": "N/A"},
                },
            },
            ("Outlook", "Banner"): {
                "inventory_targeting": "KV pub = outlook",
                "device_config": {
                    "Desktop": {"supply": "Web Only", "device_targeting": "Desktop, Tablet", "ad_sizes": "160x600, 728x90, 300x250, 300x600"},
                },
            },
            ("Outlook", "Native"): {
                "inventory_targeting": "Publisher = Outlook Native (1000230)",
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "N/A", "min_creatives": 2},
                    "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",          "ad_sizes": "N/A", "min_creatives": 2},
                    "Mobile":      {"supply": "App & Web", "device_targeting": "Mobile, Tablet",           "ad_sizes": "N/A", "min_creatives": 1},
                },
            },
            ("MCG", "Banner"): {
                "inventory_targeting": "KV pub = microsoftcasualgames",
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "300x250, 300x600, 728x90, 320x50"},
                    "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",          "ad_sizes": "300x250, 300x600, 728x90"},
                    "Mobile":      {"supply": "App & Web", "device_targeting": "Mobile, Tablet",           "ad_sizes": "320x50"},
                },
            },
            ("MCG", "Video"): {
                "inventory_targeting": "KV pub = microsoftcasualgames",
                "device_config": {
                    "Desktop": {"supply": "Web Only", "device_targeting": "Desktop, Tablet", "ad_sizes": "N/A"},
                },
            },
        },
    },
    # ── GDALI IMPRESSIONS ────────────────────────────────────────────────────
    "GDALI - Impressions": {
        "seat": "280 - Monetize",
        "line_item_type": "Guaranteed (GDALI)",
        "revenue_type": "CPM",
        "delivery_type": "Impressions",
        "allow_rtb": True,
        "underspend_catchup": "Evenly",
        "impression_pacing": "100%",
        "creative_rotation": "Evenly",
        "landing_page": "None",
        "publishers": ["MSN", "Outlook", "MCG"],
        "formats": ["Banner", "Native", "Video"],
        "priority": 5,
        "frequency_cap": "6 imps/day",
        "pacing": "Even",
        "inventory_type": "Banner",
        "geo_targeting": "Country Targeting",
        "publisher_targeting": "KV pub = msn (MSN) | KV pub = outlook (Outlook) | KV pub = microsoftcasualgames (MCG)",
        "supply_targeting": "App & Web",
        "publisher_format_config": {
            ("MSN", "Banner"): {
                "inventory_targeting": "KV pub = msn",
                "creative_specs": "300x250 OR 300x600 OR 728x90 OR 970x250 OR 320x50",
                "min_creatives": 1,
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "300x250, 300x600, 728x90, 970x250, 320x50"},
                    "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",          "ad_sizes": "728x90, 970x250, 300x250, 300x600"},
                    "Mobile":      {"supply": "Web Only",  "device_targeting": "Mobile, Tablet",           "ad_sizes": "320x50"},
                },
            },
            ("MSN", "Video"): {
                "inventory_targeting": "KV pub = msn",
                "creative_specs": "15 seconds OR 30 seconds",
                "min_creatives": 1,
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "N/A"},
                    "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",          "ad_sizes": "N/A"},
                    "Mobile":      {"supply": "Web Only",  "device_targeting": "Mobile, Tablet",           "ad_sizes": "N/A"},
                },
            },
            ("Outlook", "Banner"): {
                "inventory_targeting": "KV pub = outlook",
                "creative_specs": "160x600, 728x90, 300x250 OR 300x600",
                "min_creatives": 3,
                "device_config": {
                    "Desktop": {"supply": "Web Only", "device_targeting": "Desktop, Tablet", "ad_sizes": "728x90, 300x250, 300x600"},
                },
            },
            ("Outlook", "Native"): {
                "inventory_targeting": "Publisher = Outlook Native (1000230)",
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "N/A", "min_creatives": 2},
                    "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",          "ad_sizes": "N/A", "min_creatives": 2},
                    "Mobile":      {"supply": "App & Web", "device_targeting": "Mobile, Tablet",           "ad_sizes": "N/A", "min_creatives": 1},
                },
            },
            ("MCG", "Banner"): {
                "inventory_targeting": "KV pub = microsoftcasualgames",
                "creative_specs": "300x250, 300x600, 728x90, 320x50",
                "min_creatives": 1,
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "300x250, 300x600, 728x90, 320x50"},
                    "Desktop":     {"supply": "Web Only",  "device_targeting": "Desktop, Tablet",          "ad_sizes": "300x250, 300x600, 728x90"},
                    "Mobile":      {"supply": "App & Web", "device_targeting": "Mobile, Tablet",           "ad_sizes": "320x50"},
                },
            },
            ("MCG", "Video"): {
                "inventory_targeting": "KV pub = microsoftcasualgames",
                "creative_specs": "15 seconds OR 30 seconds",
                "min_creatives": 1,
                "device_config": {
                    "Desktop": {"supply": "Web Only", "device_targeting": "Desktop, Tablet", "ad_sizes": "N/A"},
                },
            },
        },
    },
    # ── GDALI MSN TAKEOVER ───────────────────────────────────────────────────
    "GDALI - MSN Takeover": {
        "seat": "280 - Monetize",
        "line_item_type": "Guaranteed (GDALI)",
        "revenue_type": "Fixed Fee (Cost Per Day)",
        "delivery_type": "Exclusive",
        "allow_rtb": True,
        "underspend_catchup": "Evenly",
        "impression_pacing": "100%",
        "creative_rotation": "Evenly",
        "landing_page": "None",
        "publishers": ["MSN"],
        "formats": ["Banner"],
        "priority": 15,
        "frequency_cap": "Off",
        "pacing": "Even",
        "inventory_type": "Banner",
        "geo_targeting": "Placement Targeting",
        "publisher_targeting": "Placement Targeting",
        "supply_targeting": "Web Only",
        "device_config": {
            "Desktop": {
                "supply": "Web Only", "device_targeting": "Desktop, Tablet",
                "ad_sizes": "728x90, 970x250, 300x250",
                "creative_specs": "728x90 OR 970x250 | 300x250 | 728x90",
                "min_creatives": 1, "inventory_targeting": "Placement Targeting",
            },
        },
    },
    # ── GDALI OUTLOOK TAKEOVER ───────────────────────────────────────────────
    "GDALI - Outlook Takeover": {
        "seat": "280 - Monetize",
        "line_item_type": "Guaranteed (GDALI)",
        "revenue_type": "Fixed Fee (Cost Per Day)",
        "delivery_type": "Exclusive",
        "allow_rtb": True,
        "underspend_catchup": "Evenly",
        "impression_pacing": "100%",
        "creative_rotation": "Evenly",
        "landing_page": "None",
        "publishers": ["Outlook"],
        "formats": ["Banner", "Native"],
        "priority": 15,
        "frequency_cap": "Off",
        "pacing": "Even",
        "inventory_type": "Native",
        "geo_targeting": "Country Targeting",
        "publisher_targeting": "Publisher Targeting",
        "supply_targeting": "Web Only",
        "publisher_format_config": {
            ("Outlook", "Banner"): {
                "inventory_targeting": "Publisher Targeting (Country Specific)",
                "creative_specs": "160x600, 728x90, 300x250 OR 300x600",
                "min_creatives": 3,
                "device_config": {
                    "Desktop": {"supply": "Web Only", "device_targeting": "Desktop, Tablet", "ad_sizes": "160x600, 728x90, 300x250, 300x600"},
                },
            },
            ("Outlook", "Native"): {
                "inventory_targeting": "Publisher = Outlook Native (1000230)",
                "min_creatives": 2,
                "device_config": {
                    "All Devices": {"supply": "App & Web", "device_targeting": "Desktop, Mobile, Tablet", "ad_sizes": "N/A"},
                },
            },
        },
    },
    # ── HIGH IMPACT PG ───────────────────────────────────────────────────────
    "High Impact": {
        "seat": "280 - Monetize",
        "line_item_type": "PG",
        "revenue_type": "CPM",
        "delivery_type": "N/A",
        "allow_rtb": False,
        "underspend_catchup": "N/A",
        "impression_pacing": "100%",
        "creative_rotation": "N/A",
        "landing_page": "None",
        "publishers": ["MSN"],
        "formats": ["Banner"],
        "priority": 15,
        "frequency_cap": "Off",
        "pacing": "Even",
        "inventory_type": "Banner",
        "geo_targeting": "Placement Targeting",
        "publisher_targeting": "Placement Targeting",
        "supply_targeting": "App & Web",
        "device_config": {
            "Desktop": {
                "supply": "App & Web", "device_targeting": "Desktop, Tablet",
                "ad_sizes": "ATF: 970x250, 728x90, 300x250 | BTF: 300x250, 300x600",
                "creative_specs": "ATF: 970x250 OR 728x90 OR 300x250 | BTF: 300x250 OR 300x600",
                "min_creatives": 1, "inventory_targeting": "Placement Targeting",
            },
        },
    },
}

REGION_CURRENCY = {
    "US": ("USD", "$"),
    "CA": ("CAD", "$"),
    "UK": ("GBP", "£"),
    "DE": ("EUR", "€"),
    "FR": ("EUR", "€"),
    "NL": ("EUR", "€"),
    "ES": ("EUR", "€"),
    "IT": ("EUR", "€"),
    "BE": ("EUR", "€"),
    "AT": ("EUR", "€"),
    "SE": ("EUR", "€"),
    "AU": ("AUD", "$"),
    "NZ": ("AUD", "$"),
    "SG": ("SGD", "$"),
    "IN": ("INR", "₹"),
    "JP": ("JPY", "¥"),
    "BR": ("BRL", "R$"),
    "MX": ("MXN", "$"),
}

# CPM reference ranges by product + format (align with DNV Rate Card)
CPM_RECOMMENDATIONS = {
    "PG - First Impression":        {"min": 15,   "default": 20,    "max": 40,    "note": "Premium first-in-page · MSN Banner"},
    "PG - Standard":                {"min": 8,    "default": 12,    "max": 25,    "note": "Banner/Native · MSN, Outlook, MCG"},
    "PG - Standard (Video)":        {"min": 25,   "default": 35,    "max": 65,    "note": "Video · MSN, MCG"},
    "GDALI - Impressions":          {"min": 8,    "default": 12,    "max": 25,    "note": "GDALI Guaranteed · Banner/Native"},
    "GDALI - Impressions (Video)":  {"min": 25,   "default": 35,    "max": 65,    "note": "GDALI Guaranteed · Video · MSN"},
    "GDALI - MSN Takeover":         {"min": 5000, "default": 15000, "max": 50000, "note": "Fixed Fee per day (not a CPM product)"},
    "GDALI - Outlook Takeover":     {"min": 3000, "default": 10000, "max": 30000, "note": "Fixed Fee per day (not a CPM product)"},
    "High Impact":                  {"min": 20,   "default": 30,    "max": 60,    "note": "High-impact placement · MSN Banner"},
}

# ================================================================================
# HELPER FUNCTIONS
# ================================================================================
def get_currency_symbol(market):
    """Get currency symbol based on market code."""
    return REGION_CURRENCY.get(market.upper(), ("USD", "$"))[1]

def get_compatible_formats(product, publisher):
    """Get compatible ad formats for a product + publisher combination, from the matrix."""
    product_info = PRODUCT_RULES.get(product, {})
    all_formats = product_info.get("formats", ["Banner"])

    # If product has publisher_format_config, derive available formats for this publisher
    pfc = product_info.get("publisher_format_config", {})
    if pfc:
        pub_formats = [fmt for (pub, fmt) in pfc.keys() if pub == publisher]
        if pub_formats:
            return sorted(set(pub_formats), key=["Banner", "Native", "Video"].index
                          if all(f in ["Banner", "Native", "Video"] for f in pub_formats) else str)
    return all_formats


def get_compatible_devices(product, publisher, ad_format):
    """Return valid device options constrained by the product × publisher × format device_config."""
    _order = ["All Devices", "Desktop", "Mobile", "Tablet"]
    product_info = PRODUCT_RULES.get(product, {})
    pfc = product_info.get("publisher_format_config", {})
    key = (publisher, ad_format)

    if key in pfc:
        dev_cfg = pfc[key].get("device_config", {})
    else:
        dev_cfg = product_info.get("device_config", {})

    if dev_cfg:
        devices = [d for d in _order if d in dev_cfg]
        return devices if devices else list(dev_cfg.keys())
    return ["All Devices", "Desktop", "Mobile"]


def get_product_config(product, publisher, ad_format, device="All Devices"):
    """Return a flat targeting dict for the given product / publisher / format / device."""
    base = dict(PRODUCT_RULES.get(product, {}))

    pfc = base.get("publisher_format_config", {})
    key = (publisher, ad_format)

    if key in pfc:
        fmt_cfg = dict(pfc[key])
        dev_cfg = fmt_cfg.pop("device_config", {})
        fallback = dev_cfg.get("All Devices", next(iter(dev_cfg.values()), {}) if dev_cfg else {})
        dev_specific = dict(dev_cfg.get(device, fallback))
        result = {**base, **fmt_cfg, **dev_specific}
    else:
        dev_cfg = base.get("device_config", {})
        fallback = dev_cfg.get("All Devices", next(iter(dev_cfg.values()), {}) if dev_cfg else {})
        dev_specific = dict(dev_cfg.get(device, fallback))
        result = {**base, **dev_specific}

    for k in ("publisher_format_config", "device_config"):
        result.pop(k, None)
    return result

def generate_o_o_taxonomy(market, publisher, product, ad_format, device):
    """Generate O&O Naming Taxonomy — uses Views suffix for video, Imps for display."""
    device_code = {"All Devices": "AllDevices", "Desktop": "Desktop", "Mobile": "Mobile", "Tablet": "Tablet"}.get(device, "AllDevices")
    metric = "Views" if ad_format == "Video" else "Imps"
    li_code = PRODUCT_RULES.get(product, {}).get("line_item_type", "PG").replace("Guaranteed (", "").replace(")", "")
    return f"{market}_{publisher}_{product}_{ad_format}_{device_code}_{li_code}_{metric}"

def generate_li_name(advertiser, market, publisher, product):
    """Generate Line Item Name."""
    return f"{advertiser}_{market}_{publisher}_{product}"

def is_video_product(product_name, ad_format=None):
    """Return True if the current format is Video."""
    if ad_format is not None:
        return ad_format == "Video"
    return PRODUCT_RULES.get(product_name, {}).get("inventory_type") == "Video"

def get_default_cpm(product_name, ad_format=None):
    """Return the suggested default CPM for a product, adjusted for video."""
    if ad_format == "Video":
        video_key = f"{product_name} (Video)"
        if video_key in CPM_RECOMMENDATIONS:
            return float(CPM_RECOMMENDATIONS[video_key]["default"])
    return float(CPM_RECOMMENDATIONS.get(product_name, {"default": 10.0})["default"])

# ── DEAL-TYPE / FORMAT RESOLUTION ────────────────────────────────────────────

# Spec §4.3 — allowed formats per deal type
_FORMAT_ALLOWLIST = {
    "GDALI - Impressions":      ["Banner", "Video", "Native"],
    "GDALI - MSN Takeover":     ["Banner"],
    "GDALI - Outlook Takeover": ["Banner", "Native"],
    "PG - Standard":            ["Banner", "Video", "Native"],
    "PG - First Impression":    ["Banner"],
    "High Impact":              ["Banner"],
}

def is_cpd_product(product: str) -> bool:
    """True for Fixed Fee (Cost Per Day) products — Takeovers."""
    return PRODUCT_RULES.get(product, {}).get("revenue_type", "") == "Fixed Fee (Cost Per Day)"

def resolve_inventory_type(ad_format: str) -> str:
    """Spec §4.1: Video/Native must be App & Web; Banner is Web Only."""
    return "App & Web" if ad_format in ("Video", "Native") else "Web Only"

def resolve_inventory_targeting_mode(publisher: str, ad_format: str) -> str:
    """Spec §4.2: KV targeting for MSN/MCG; Publisher ID for Outlook Native."""
    if publisher == "Outlook" and ad_format == "Native":
        return "Publisher ID — Outlook Native (1000230)"
    mapping = {
        "MSN":     "Key/Value — pub = msn",
        "Outlook": "Key/Value — pub = outlook",
        "MCG":     "Key/Value — pub = microsoftcasualgames",
    }
    return mapping.get(publisher, f"Key/Value — pub = {publisher.lower()}")

def validate_deal_format(product: str, ad_format: str) -> tuple[bool, str]:
    """Spec §4.3: Return (valid, reason). One format per line item; format must be allowed."""
    allowed = _FORMAT_ALLOWLIST.get(product, ["Banner", "Video", "Native"])
    if ad_format not in allowed:
        return False, f"{ad_format} is not allowed for {product} (allowed: {', '.join(allowed)})"
    return True, ""

# Spec §4.4 — canonical banner sizes per publisher × device
_BANNER_SIZES = {
    ("MSN",     "All Devices"): ["300x250", "300x600", "728x90", "970x250", "320x50"],
    ("MSN",     "Desktop"):     ["970x250", "728x90", "300x250", "300x600"],
    ("MSN",     "Mobile"):      ["320x50"],
    ("Outlook", "Desktop"):     ["160x600", "728x90", "300x250", "300x600"],
    ("MCG",     "All Devices"): ["300x250", "300x600", "728x90", "320x50"],
    ("MCG",     "Desktop"):     ["300x250", "300x600", "728x90"],
    ("MCG",     "Mobile"):      ["320x50"],
}

def get_allowed_banner_sizes(publisher: str, device: str) -> list:
    """Spec §4.4: Canonical sizes for publisher × device. Returns list of size strings."""
    return _BANNER_SIZES.get((publisher, device), _BANNER_SIZES.get((publisher, "All Devices"), []))

def get_freq_cap_numeric(product: str) -> int:
    """Extract numeric frequency cap from product rules. Returns 0 if uncapped/CPD."""
    cap_text = PRODUCT_RULES.get(product, {}).get("frequency_cap", "")
    if cap_text in ("Off", "N/A", "") or is_cpd_product(product):
        return 0
    if "1 imp" in cap_text:
        return 1
    if "6 imp" in cap_text:
        return 6
    return 6  # safe default


# ── FORECASTING FORMULAS (Spec §6) ───────────────────────────────────────────

def calc_avg_daily_uniques(total_impressions: float, flight_days: int,
                           freq_cap_num: int) -> float:
    """
    Spec §6.3: Avg Uniques (Daily) at product level.
    Estimate = daily_impressions / freq_cap  (users × exposures/user = impressions).
    When freq_cap is 0 (uncapped/CPD) returns 0 (not applicable).
    """
    if flight_days <= 0 or freq_cap_num <= 0:
        return 0.0
    daily_imps = total_impressions / flight_days
    return daily_imps / freq_cap_num

def calc_avg_frequency_daily(total_impressions: float, avg_daily_uniques: float,
                             flight_days: int) -> float:
    """
    Spec §6.1: Avg Frequency (Daily) = daily_impressions / daily_uniques.
    Returns 0 when not applicable (CPD / uncapped / no uniques).
    """
    if flight_days <= 0 or avg_daily_uniques <= 0:
        return 0.0
    return (total_impressions / flight_days) / avg_daily_uniques

def calc_lifetime_uniques(avg_daily_uniques: float, turnover_per_day: float,
                          days_lifetime: int) -> float:
    """
    Spec §6.4: Lifetime Uniques = avg_daily_uniques × (1 + turnover × (days - 1)).
    Not applicable for CPD products (caller must guard).
    """
    if avg_daily_uniques <= 0 or days_lifetime <= 0:
        return 0.0
    return avg_daily_uniques * (1 + turnover_per_day * (days_lifetime - 1))


def calculate_delivery_pressure(flights_list, product_name, frequency_cap_text=""):
    """
    Calculate delivery success pressure: Low / Medium / High.
    
    Logic:
    - LOW: Duration ≥ 30 days AND Daily Volume ≤ Freq Cap
    - MEDIUM: Duration 14-29 days OR Daily Volume manageable
    - HIGH: Duration < 14 days AND High daily volume, or very constrained
    """
    if not flights_list:
        return "Low", "No flights configured."
    
    total_volume = sum(f.get("volume", 0) for f in flights_list)
    # Get flight duration from first flight (assume all same product)
    first_flight = flights_list[0]
    start = first_flight.get("start_date")
    end = first_flight.get("end_date")
    
    if isinstance(start, str):
        from datetime import datetime
        start = datetime.strptime(start, "%Y-%m-%d").date()
    if isinstance(end, str):
        from datetime import datetime
        end = datetime.strptime(end, "%Y-%m-%d").date()
    
    flight_days = max(1, (end - start).days + 1)
    daily_volume = total_volume / flight_days if flight_days > 0 else total_volume
    
    # Frequency cap extraction (rough: "6 impressions per user per day" -> 6)
    freq_cap = 6  # default
    if "1 impression" in frequency_cap_text.lower():
        freq_cap = 1
    elif "2 impression" in frequency_cap_text.lower():
        freq_cap = 2
    
    # Pressure index
    pressure_ratio = (daily_volume / freq_cap) if freq_cap > 0 else 0
    
    if flight_days >= 30 and pressure_ratio <= 1.0:
        return "Low", f"Flight duration {flight_days} days. Daily volume manageable. Low activation risk."
    elif flight_days >= 14 and pressure_ratio <= 1.5:
        return "Medium", f"Flight duration {flight_days} days. Moderate delivery pressure. Confirm frequency caps and date windows."
    else:
        return "High", f"Flight duration {flight_days} days. High daily volume vs. frequency cap. Escalate to ops team."

def create_internal_excel_export(flights_list, campaign_info, delivery_pressure_label,
                                 product_config, unified_checklist):
    """Create internal-facing Excel with targeting setup and QC checklist,
    fully aligned with the on-screen Risk & Quality Control section."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Setup"

    section_font = Font(bold=True, size=11, color="0078D4")
    header_fill  = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font  = Font(bold=True)

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "Media Plan - Internal Setup & QC"
    title.font = Font(bold=True, size=14, color="0078D4")

    # ── Campaign details ─────────────────────────────────────────────────────
    row = 3
    for label, key in [("Campaign Name:", "campaign"), ("Advertiser:", "advertiser"),
                        ("Market:", "market"), ("Publisher:", "publisher"),
                        ("Product:", "product")]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = campaign_info.get(key, "")
        row += 1

    # ── INTERNAL SETUP & TARGETING (product-specific) ────────────────────────
    row += 1
    ws[f"A{row}"] = "INTERNAL SETUP & TARGETING"
    ws[f"A{row}"].font = section_font
    row += 2

    for col, header in enumerate(["Field", "Value"], 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    setup_fields = [
        ("Seat ID",              product_config.get("seat", "280 - Monetize")),
        ("Line Item Type",       product_config.get("line_item_type", "N/A")),
        ("Revenue Type",         product_config.get("revenue_type", "CPM")),
        ("Priority",             product_config.get("priority", "N/A")),
        ("Frequency Cap",        product_config.get("frequency_cap", "N/A")),
        ("Pacing",               product_config.get("pacing", "N/A")),
        ("Inventory Type",       product_config.get("inventory_type",
                                    resolve_inventory_type(campaign_info.get("format", "Banner")))),
        ("Inventory Targeting",  product_config.get("inventory_targeting",
                                    product_config.get("publisher_targeting", "N/A"))),
        ("Supply",               product_config.get("supply",
                                    product_config.get("supply_targeting", "N/A"))),
        ("Ad Sizes",             product_config.get("ad_sizes", "N/A")),
        ("Device Targeting",     product_config.get("device_targeting", "N/A")),
        ("Creative Specs",       product_config.get("creative_specs", "N/A")),
        ("Allow RTB",            str(product_config.get("allow_rtb", "N/A"))),
        ("Underspend Catchup",   product_config.get("underspend_catchup", "N/A")),
        ("Geo Targeting",        product_config.get("geo_targeting", "Country Targeting")),
    ]

    for field_name, field_value in setup_fields:
        ws.cell(row=row, column=1).value = field_name
        ws.cell(row=row, column=2).value = str(field_value)
        row += 1

    # ── QC CHECKLIST (mirrors unified_checklist from the UI) ─────────────────
    row += 2
    ws[f"A{row}"] = "QC CHECKLIST"
    ws[f"A{row}"].font = section_font
    row += 2

    for col, header in enumerate(["Checklist Item", "Status / Value"], 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    for item_name, item_value in unified_checklist:
        ws.cell(row=row, column=1).value = str(item_name)
        ws.cell(row=row, column=2).value = str(item_value)
        row += 1

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 45

    return wb

def create_excel_export(flights_list, campaign_info, delivery_pressure_label, product_config, cpm_rate, currency_sym, vcr_target=None, completed_views=None):
    """Create IO Excel file with Microsoft-style formatting and forecasting table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Plan"
    
    # Colors
    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10, color="0078D4")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "MEDIA PLAN - MONETIZE GUARANTEED"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Campaign Details Header
    row = 3
    ws[f"A{row}"] = "Advertiser:"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws[f"B{row}"] = campaign_info.get("advertiser", "")
    ws[f"B{row}"].font = Font(size=11)
    
    ws[f"D{row}"] = "Market:"
    ws[f"D{row}"].font = Font(bold=True, size=11)
    ws[f"E{row}"] = campaign_info.get("market", "")
    ws[f"E{row}"].font = Font(size=11)
    
    row = 4
    ws[f"A{row}"] = "Campaign:"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws[f"B{row}"] = campaign_info.get("campaign", "")
    ws[f"B{row}"].font = Font(size=11)
    
    ws[f"D{row}"] = "Publisher:"
    ws[f"D{row}"].font = Font(bold=True, size=11)
    ws[f"E{row}"] = campaign_info.get("publisher", "")
    ws[f"E{row}"].font = Font(size=11)
    
    row = 5
    ws[f"A{row}"] = "Order Type:"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws[f"B{row}"] = "Insertion Order"
    ws[f"B{row}"].font = Font(size=11)
    
    ws[f"D{row}"] = "Request Date:"
    ws[f"D{row}"].font = Font(bold=True, size=11)
    ws[f"E{row}"] = datetime.now().strftime("%m/%d/%Y")
    ws[f"E{row}"].font = Font(size=11)
    
    # Campaign Overview
    row = 7
    ws[f"A{row}"] = "CAMPAIGN OVERVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="0078D4")

    row = 8
    overview_headers = ["#", "Product", "Publisher", "Format", "Start Date", "End Date", "Budget"]
    for col, header in enumerate(overview_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_budget = 0
    for row_idx, flight in enumerate(flights_list, 9):
        flight_budget = flight.get('budget', 0)
        total_budget += flight_budget

        ws.cell(row=row_idx, column=1).value = flight.get("flight_num", 1)
        ws.cell(row=row_idx, column=2).value = flight.get("product", "")
        ws.cell(row=row_idx, column=3).value = flight.get("publisher", "")
        ws.cell(row=row_idx, column=4).value = flight.get("format", "")

        start_date = flight.get("start_date")
        ws.cell(row=row_idx, column=5).value = start_date.strftime("%m/%d/%Y") if hasattr(start_date, "strftime") else str(start_date)
        end_date = flight.get("end_date")
        ws.cell(row=row_idx, column=6).value = end_date.strftime("%m/%d/%Y") if hasattr(end_date, "strftime") else str(end_date)
        ws.cell(row=row_idx, column=7).value = f"{flight_budget:,.2f}"
    
    # Per-product Forecasting Table
    forecast_row = len(flights_list) + 11
    ws[f"A{forecast_row}"] = "FORECASTING BY PRODUCT"
    ws[f"A{forecast_row}"].font = Font(bold=True, size=11, color="0078D4")

    forecast_row += 1
    f_headers = ["Product", "Publisher", "Format", "CPM", "Budget", "Total Imps", "Est. Daily Avg"]
    for col, header in enumerate(f_headers, 1):
        cell = ws.cell(row=forecast_row, column=col)
        cell.value = header
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    forecast_row += 1
    # Group flights by product/publisher/format
    fgroups = {}
    for fl in flights_list:
        fkey = (fl.get("product", ""), fl.get("publisher", ""), fl.get("format", ""))
        if fkey not in fgroups:
            fgroups[fkey] = {"budget": 0, "imps": 0, "cpm": fl.get("cpm", cpm_rate), "days": 0}
        fgroups[fkey]["budget"] += fl.get("budget", 0)
        fgroups[fkey]["imps"] += fl.get("volume", 0)
        sd, ed = fl.get("start_date"), fl.get("end_date")
        if hasattr(sd, "__sub__") and hasattr(ed, "__sub__"):
            fgroups[fkey]["days"] += (ed - sd).days + 1

    grand_budget = grand_imps = 0
    for (fprod, fpub, ffmt), fd in fgroups.items():
        _is_cpd_fl = is_cpd_product(fprod)
        if _is_cpd_fl:
            row_data = [fprod, fpub, ffmt,
                        f"Fixed Fee (CPD)",
                        f"{fd['budget']:,.2f} {currency_sym}",
                        f"{fd['days']} day(s)",
                        f"{fd['budget']/max(1,fd['days']):,.2f} {currency_sym}/day"]
        else:
            daily = fd["imps"] / max(1, fd["days"])
            row_data = [fprod, fpub, ffmt,
                        f"{fd['cpm']:,.2f} {currency_sym}",
                        f"{fd['budget']:,.2f} {currency_sym}",
                        f"{fd['imps']:,.0f}",
                        f"{daily:,.0f}"]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=forecast_row, column=col).value = val
        grand_budget += fd["budget"]
        grand_imps += fd["imps"]
        forecast_row += 1

    if len(fgroups) > 1:
        total_row = ["TOTAL", "", "", "—", f"{grand_budget:,.2f} {currency_sym}", f"{grand_imps:,.0f}", "—"]
        for col, val in enumerate(total_row, 1):
            c = ws.cell(row=forecast_row, column=col)
            c.value = val
            c.font = Font(bold=True)
        forecast_row += 1

    # Video metrics
    if vcr_target is not None:
        ws.cell(row=forecast_row, column=1).value = "Target VCR (%)"
        ws.cell(row=forecast_row, column=1).font = Font(bold=True)
        ws.cell(row=forecast_row, column=2).value = f"{vcr_target}%"
        forecast_row += 1
        ws.cell(row=forecast_row, column=1).value = "Est. Completed Views"
        ws.cell(row=forecast_row, column=1).font = Font(bold=True)
        ws.cell(row=forecast_row, column=2).value = f"{completed_views:,.0f}" if completed_views else "—"
        forecast_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    
    return wb

def create_pdf_export(flights_list, campaign_info, delivery_pressure_label, product_config, cpm_rate, currency_sym, vcr_target=None, completed_views=None):
    """Create IO PDF file with same information as Excel."""
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0078D4'),
        spaceAfter=20,
        alignment=1  # center
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#0078D4'),
        spaceAfter=10,
        spaceBefore=10
    )
    
    # Title
    elements.append(Paragraph("MEDIA PLAN - MONETIZE GUARANTEED", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Campaign Details
    campaign_data = [
        ["Advertiser:", campaign_info.get("advertiser", ""), "Market:", campaign_info.get("market", "")],
        ["Campaign:", campaign_info.get("campaign", ""), "Publisher:", campaign_info.get("publisher", "")],
        ["Order Type:", "Insertion Order", "Request Date:", datetime.now().strftime("%m/%d/%Y")],
    ]
    
    campaign_table = Table(campaign_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    campaign_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D9E1F2')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#D9E1F2')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    elements.append(campaign_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Campaign Overview
    elements.append(Paragraph("CAMPAIGN OVERVIEW", heading_style))

    overview_data = [["#", "Product", "Publisher", "Format", "Start Date", "End Date", "Budget"]]
    total_budget = 0

    for flight in flights_list:
        flight_budget = flight.get('budget', 0)
        total_budget += flight_budget
        start_date = flight.get("start_date")
        start_str = start_date.strftime("%m/%d/%Y") if hasattr(start_date, "strftime") else str(start_date)
        end_date = flight.get("end_date")
        end_str = end_date.strftime("%m/%d/%Y") if hasattr(end_date, "strftime") else str(end_date)
        overview_data.append([
            str(flight.get("flight_num", 1)),
            flight.get("product", ""),
            flight.get("publisher", ""),
            flight.get("format", ""),
            start_str, end_str,
            f"{flight_budget:,.2f}"
        ])

    overview_data.append(["", "", "", "", "", "TOTAL:", f"{total_budget:,.2f} {currency_sym}"])

    overview_table = Table(overview_data, colWidths=[0.4*inch, 1.2*inch, 0.8*inch, 0.7*inch, 1.1*inch, 1.1*inch, 1.0*inch])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(overview_table)
    elements.append(Spacer(1, 0.3*inch))

    # Per-product Forecasting
    elements.append(Paragraph("FORECASTING BY PRODUCT", heading_style))

    fgroups = {}
    for fl in flights_list:
        fkey = (fl.get("product", ""), fl.get("publisher", ""), fl.get("format", ""))
        if fkey not in fgroups:
            fgroups[fkey] = {"budget": 0, "imps": 0, "cpm": fl.get("cpm", cpm_rate), "days": 0}
        fgroups[fkey]["budget"] += fl.get("budget", 0)
        fgroups[fkey]["imps"] += fl.get("volume", 0)
        sd, ed = fl.get("start_date"), fl.get("end_date")
        if hasattr(sd, "__sub__") and hasattr(ed, "__sub__"):
            fgroups[fkey]["days"] += (ed - sd).days + 1

    forecast_data = [["Product", "Publisher", "Format", "CPM", "Budget", "Total Imps", "Daily Avg"]]
    grand_budget = grand_imps = 0
    for (fprod, fpub, ffmt), fd in fgroups.items():
        _is_cpd_fl = is_cpd_product(fprod)
        if _is_cpd_fl:
            forecast_data.append([fprod, fpub, ffmt,
                                   "Fixed Fee (CPD)",
                                   f"{fd['budget']:,.2f} {currency_sym}",
                                   f"{fd['days']} day(s)",
                                   f"{fd['budget']/max(1,fd['days']):,.2f} {currency_sym}/day"])
        else:
            daily = fd["imps"] / max(1, fd["days"])
            forecast_data.append([fprod, fpub, ffmt,
                                   f"{fd['cpm']:,.2f} {currency_sym}",
                                   f"{fd['budget']:,.2f} {currency_sym}",
                                   f"{fd['imps']:,.0f}",
                                   f"{daily:,.0f}"])
        grand_budget += fd["budget"]
        grand_imps += fd["imps"]

    if len(fgroups) > 1:
        forecast_data.append(["TOTAL", "", "", "—",
                               f"{grand_budget:,.2f} {currency_sym}",
                               "—", "—"])

    if vcr_target is not None:
        forecast_data.append(["Target VCR (%)", f"{vcr_target}%", "", "", "", "", ""])
        forecast_data.append(["Est. Completed Views", f"{completed_views:,.0f}" if completed_views else "—", "", "", "", "", ""])
    
    forecast_table = Table(forecast_data, colWidths=[1.3*inch, 0.9*inch, 0.7*inch, 0.9*inch, 1.1*inch, 0.9*inch, 0.9*inch])
    forecast_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(forecast_table)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

# ================================================================================
# 2. CAMPAIGN DETAILS (INPUTS)
# ================================================================================
st.markdown("## Campaign Details")
st.markdown('<div style="margin-bottom: 1.5rem;"></div>', unsafe_allow_html=True)

col1, col2, col3, col4 = st.columns(4)

with col1:
    market = st.selectbox(
        "Market",
        list(REGION_CURRENCY.keys()),
        key="market",
        help="Country/Region for campaign"
    )
    st.session_state.campaign_details["market"] = market

with col2:
    advertiser = st.text_input(
        "Advertiser",
        value=st.session_state.campaign_details.get("advertiser", ""),
        key="advertiser",
        help="Client/Advertiser name"
    )
    st.session_state.campaign_details["advertiser"] = advertiser

with col3:
    campaign = st.text_input(
        "Campaign",
        value=st.session_state.campaign_details.get("campaign", ""),
        key="campaign",
        help="Campaign name/ID"
    )
    st.session_state.campaign_details["campaign"] = campaign

with col4:
    publisher = st.selectbox(
        "Publisher",
        ["MSN", "Outlook", "MCG"],
        key="publisher",
        help="Target publisher (MSN, Outlook, MCG = Microsoft Casual Games)"
    )
    st.session_state.campaign_details["publisher"] = publisher

col5, col6, col7, col8 = st.columns(4)

with col5:
    _tmp_product = st.session_state.get("product", list(PRODUCT_RULES.keys())[0])
    _tmp_publisher = st.session_state.get("publisher", "MSN")
    _tmp_format = st.session_state.get("format", "Banner")
    _compatible_devices = get_compatible_devices(_tmp_product, _tmp_publisher, _tmp_format)
    device = st.selectbox(
        "Device",
        _compatible_devices,
        key="device",
        help="Device options filtered by product · publisher · format"
    )
    st.session_state.campaign_details["device"] = device

with col6:
    product = st.selectbox(
        "Product",
        list(PRODUCT_RULES.keys()),
        key="product",
        help="Monetize Guaranteed product"
    )
    st.session_state.campaign_details["product"] = product

with col7:
    ad_format = st.selectbox(
        "Format",
        get_compatible_formats(product, publisher),
        key="format",
        help="Ad format (locked by product/publisher)"
    )
    st.session_state.campaign_details["format"] = ad_format

with col8:
    st.write("")

# ── Inline config validation (Spec §4.2–4.3) ─────────────────────────────────
_early_fmt_valid, _early_fmt_err = validate_deal_format(product, ad_format)
if not _early_fmt_valid:
    st.error(f"❌ **Format conflict:** {_early_fmt_err}")
if publisher == "Outlook" and ad_format == "Video":
    st.error("❌ **Outlook Video does not exist.** Select Banner or Native for Outlook.")
_early_inv_type = resolve_inventory_type(ad_format)
_early_inv_mode = resolve_inventory_targeting_mode(publisher, ad_format)
st.caption(
    f"🔒 Inventory: **{_early_inv_type}** · Targeting Mode: **{_early_inv_mode}** "
    f"· Format enforcement: {'✅' if _early_fmt_valid else '❌'}"
)

col9, col10 = st.columns(2)
with col9:
    start_date = st.date_input(
        "Start Date",
        value=datetime.now().date(),
        key="start_date"
    )
    st.session_state.campaign_details["start_date"] = start_date

with col10:
    end_date = st.date_input(
        "End Date",
        value=datetime.now().date() + timedelta(days=30),
        key="end_date"
    )
    st.session_state.campaign_details["end_date"] = end_date

# ================================================================================
# 3. O&O NAMING TAXONOMY
# ================================================================================
st.markdown("## O&O Naming Taxonomy")
st.markdown('<div style="margin-bottom: 1.5rem;"></div>', unsafe_allow_html=True)

taxonomy = generate_o_o_taxonomy(market, publisher, product, ad_format, device)
li_name = generate_li_name(advertiser or "ADVERTISER", market, publisher, product)

col_tax1, col_tax2 = st.columns(2)

with col_tax1:
    st.markdown("**O&O Taxonomy**")
    st.code(taxonomy, language="text")

with col_tax2:
    st.markdown("**Line Item Name**")
    st.code(li_name, language="text")

# ================================================================================
# 4. PRICING
# ================================================================================
st.markdown("## Pricing")
st.markdown('<div style="margin-bottom: 1.5rem;"></div>', unsafe_allow_html=True)

currency_sym = get_currency_symbol(market)
currency_code = REGION_CURRENCY.get(market.upper(), ("USD", "$"))[0]

# ── Pricing varies by product type: CPD for Takeovers, CPM for everything else ──
_takeover = is_cpd_product(product)

# Initialize video defaults (may be overwritten below)
vcr_target = 75
completed_views = 0
video_duration = None
cpm_rate = 0.0
impressions = 0
budget = 0.0
_cpd_rate = 0.0
_flight_days_pricing = max(1, (end_date - start_date).days + 1)

if _takeover:
    # ── CPD mode ─────────────────────────────────────────────────────────────
    st.markdown("**Forecasting (CPD):** Total Cost = Daily Rate × Number of Days")
    _rec = CPM_RECOMMENDATIONS.get(product, {"default": 10000, "min": 3000, "max": 50000, "note": "Fixed Fee per day"})
    col_price1, col_price2, col_price3 = st.columns(3)
    with col_price1:
        _cpd_rate = st.number_input(
            f"Daily Rate / Cost Per Day ({currency_sym})",
            min_value=0.0,
            value=float(_rec["default"]),
            step=100.0,
            key="budget_value",
            help=f"{_rec.get('note','')} · Align with DNV Rate Card"
        )
    with col_price2:
        st.metric("Flight Duration", f"{_flight_days_pricing} days")
    with col_price3:
        total_cpd_cost = _cpd_rate * _flight_days_pricing
        st.metric("Total Estimated Cost", f"{currency_sym}{total_cpd_cost:,.2f}")
    budget = total_cpd_cost
    cpm_rate = 0.0
    impressions = 0.0
    st.info(
        f"📅 **{product}** · Fixed Fee · "
        f"{currency_sym}{_cpd_rate:,.2f}/day × {_flight_days_pricing} days = "
        f"**{currency_sym}{total_cpd_cost:,.2f} total**  ·  Confirm rate with DNV Rate Card for {market}."
    )
else:
    # ── CPM mode ─────────────────────────────────────────────────────────────
    st.markdown("**Forecasting:** Impressions = (Spend ÷ CPM) × 1,000")
    _cpm_key = f"{product} (Video)" if ad_format == "Video" and f"{product} (Video)" in CPM_RECOMMENDATIONS else product
    _cpm_rec = CPM_RECOMMENDATIONS.get(_cpm_key, {"default": get_default_cpm(product, ad_format), "min": 5.0, "max": 200.0, "note": ""})
    col_price1, col_price2, col_price3 = st.columns(3)
    with col_price1:
        budget = st.number_input(
            f"Budget/Spend ({currency_sym})",
            min_value=0.0,
            value=1000.0,
            step=0.01,
            key="budget_value"
        )
    with col_price2:
        cpm_rate = st.number_input(
            f"CPM ({currency_sym})",
            min_value=0.01,
            value=float(_cpm_rec["default"]),
            step=0.01,
            key="cpm_rate",
            help=f"{_cpm_rec.get('note', '')} · Align with DNV Rate Card"
        )
    with col_price3:
        impressions = (budget / cpm_rate) * 1000 if cpm_rate > 0 else 0.0
        st.metric("Calculated Impressions", f"{impressions:,.0f}")
    st.caption(f"💡 Align CPM {currency_sym}{cpm_rate:.2f} with the DNV Rate Card for **{market}** · {_cpm_rec.get('note', '')}")

total_cost = budget

# Video-specific metrics (CPM products only)
if not _takeover and is_video_product(product, ad_format):
    st.markdown("**🎥 Video Metrics**")
    _dur_col, _ = st.columns([1, 3])
    with _dur_col:
        video_duration = st.selectbox(
            "Creative Duration", ["15 seconds", "30 seconds"], key="video_duration",
            help="Video creative length · affects creative specs and delivery setup"
        )
    col_vid1, col_vid2, col_vid3 = st.columns(3)
    with col_vid1:
        vcr_target = st.slider(
            "Target VCR (%)", min_value=50, max_value=100, value=75, step=5,
            help="Video Completion Rate — % of impressions watched to completion"
        )
    with col_vid2:
        completed_views = impressions * (vcr_target / 100)
        st.metric("Estimated Completed Views", f"{completed_views:,.0f}")
    with col_vid3:
        if completed_views > 0:
            cost_per_view = budget / completed_views
            st.metric("Est. Cost / Completed View", f"{currency_sym}{cost_per_view:.4f}")
        else:
            st.metric("Est. Cost / Completed View", "—")

# ── AUDIENCE & FREQUENCY ESTIMATION (Spec §6) ────────────────────────────────
_is_cpd = is_cpd_product(product)
_freq_cap_num = get_freq_cap_numeric(product)

# Pre-compute base metrics for current flight window
_flight_days_est = max(1, (end_date - start_date).days + 1)
_avg_daily_uniques_est = calc_avg_daily_uniques(impressions, _flight_days_est, _freq_cap_num)
_avg_freq_est          = calc_avg_frequency_daily(impressions, _avg_daily_uniques_est, _flight_days_est)

# Turnover rate slider — only when forecasting is applicable
_turnover_rate = 0.0
_lifetime_uniques_est = 0.0
if not _is_cpd and _freq_cap_num > 0:
    st.markdown("**Audience Reach Forecast** _(applicable only to CPM products)_")
    _aud_col1, _aud_col2, _aud_col3, _aud_col4 = st.columns(4)
    with _aud_col1:
        _turnover_rate = st.slider(
            "Daily Audience Turnover (%)", min_value=5, max_value=50,
            value=30, step=5,
            help="% of unique users refreshing per day — 25–35 % is typical for MSN/Outlook",
            key="turnover_rate",
        ) / 100.0
    with _aud_col2:
        st.metric("Est. Avg Daily Uniques", f"{_avg_daily_uniques_est:,.0f}",
                  help="daily_impressions ÷ freq_cap")
    with _aud_col3:
        st.metric("Est. Avg Daily Frequency", f"{_avg_freq_est:.2f}",
                  help="daily_impressions ÷ daily_uniques")
    with _aud_col4:
        _lifetime_uniques_est = calc_lifetime_uniques(
            _avg_daily_uniques_est, _turnover_rate, _flight_days_est
        )
        st.metric("Est. Lifetime Uniques", f"{_lifetime_uniques_est:,.0f}",
                  help="avg_daily_uniques × (1 + turnover × (days − 1))")
elif _is_cpd:
    # ── CPD summary panel ──────────────────────────────────────────────────
    st.markdown("**📅 CPD Delivery Forecast** _(Fixed Fee — time-based delivery)_")
    _cpd_col1, _cpd_col2, _cpd_col3 = st.columns(3)
    with _cpd_col1:
        st.metric("Daily Rate", f"{currency_sym}{_cpd_rate:,.2f}")
    with _cpd_col2:
        st.metric("Takeover Days", f"{_flight_days_est}")
    with _cpd_col3:
        st.metric("Total Fee", f"{currency_sym}{_cpd_rate * _flight_days_est:,.2f}")
    st.caption("📌 Impression frequency metrics are **not applicable** for CPD products — SOV delivery is exclusive and time-gated.")

# ================================================================================
# 5. TARGETING RULES (LOCKED BY PRODUCT)
# ================================================================================
st.markdown('<div style="margin-bottom: 1.5rem;"></div>', unsafe_allow_html=True)
st.markdown("## Targeting Rules")

product_config = get_product_config(product, publisher, ad_format, device)

st.markdown("**Product Configuration** _(Locked by product × publisher × format × device)_")

col_rule1, col_rule2, col_rule3, col_rule4 = st.columns(4)

with col_rule1:
    st.metric("Seat", product_config.get("seat", "280"))

with col_rule2:
    st.metric("Priority", product_config.get("priority", "-"))

with col_rule3:
    st.metric("Pacing", product_config.get("pacing", "Even"))

with col_rule4:
    st.metric("Revenue Type", product_config.get("revenue_type", "CPM"))

rules_data = {
    "Frequency Cap":      product_config.get("frequency_cap", "N/A"),
    "Inventory Targeting": product_config.get("inventory_targeting",
                           product_config.get("publisher_targeting", "N/A")),
    "Supply":             product_config.get("supply", product_config.get("supply_targeting", "N/A")),
    "Ad Sizes":           product_config.get("ad_sizes", "N/A"),
    "Device Targeting":   product_config.get("device_targeting", "N/A"),
    "Geo Targeting":      product_config.get("geo_targeting", "N/A"),
    "Line Item Type":     product_config.get("line_item_type", "N/A"),
    "Creative Specs":     product_config.get("creative_specs", "N/A"),
    "Allow RTB":          str(product_config.get("allow_rtb", "N/A")),
    "Underspend Catchup": product_config.get("underspend_catchup", "N/A"),
}

st.markdown("**Targeting Details:**")
for key, value in rules_data.items():
    st.write(f"• **{key}:** {value}")

# ================================================================================
# 6. MONETIZE SETUP (COMPACT CHECKLIST)
# ================================================================================
st.markdown('<div style="margin-bottom: 1rem;"></div>', unsafe_allow_html=True)
st.markdown("## Monetize Setup")

product_rule_str = (
    f"Seat 280 \u00b7 {product_config.get('line_item_type', 'PG')} \u00b7 Priority {product_config.get('priority', '-')} "
    f"\u00b7 {product_config.get('pacing', 'Even')} pacing \u00b7 Freq Cap: {product_config.get('frequency_cap', 'N/A')} "
    f"\u00b7 Targeting: {product_config.get('inventory_targeting', product_config.get('publisher_targeting', 'N/A'))} "
    f"\u00b7 Ad Sizes: {product_config.get('ad_sizes', 'N/A')} \u00b7 Format: {ad_format} \u00b7 {market}"
)
st.info(product_rule_str)

# ================================================================================
# 7. MANAGE MULTIPLE FLIGHTS
# ================================================================================
st.markdown('<div style="margin-bottom: 1rem;"></div>', unsafe_allow_html=True)
st.markdown("## Multiple Flights")

st.write(f"**Current Flights:** {len(st.session_state.flights)}")

if st.button("➕ Add New Flight (Same Product, Different Dates & Budget)", use_container_width=True):
    new_flight = {
        "flight_num": len(st.session_state.flights) + 1,
        "product": product,
        "publisher": publisher,
        "format": ad_format,
        "device": device,
        "start_date": start_date,
        "end_date": end_date,
        "budget": budget,
        "cpm": cpm_rate,
        "volume": impressions,
        "currency": currency_sym,
        "total_cost": total_cost,
        "duration": video_duration,
    }
    st.session_state.flights.append(new_flight)
    st.success(f"Flight {new_flight['flight_num']} added — {product} | {publisher} {ad_format} | {start_date} to {end_date}")

if st.session_state.flights:
    st.markdown("**Existing Flights:**")
    for idx, flight in enumerate(st.session_state.flights):
        col_flight1, col_flight2, col_flight3 = st.columns([3, 1, 1])
        
        with col_flight1:
            pub_label = flight.get('publisher', publisher)
            fmt_label = flight.get('format', ad_format)
            st.write(f"**Flight {flight['flight_num']}** | {flight['product']} | {pub_label} {fmt_label} | {flight['start_date']} → {flight['end_date']}")
        
        with col_flight2:
            st.write(f"{flight['currency']}{flight['total_cost']:,.2f}")
        
        with col_flight3:
            if st.button("🗑️ Remove", key=f"remove_flight_{idx}"):
                st.session_state.flights.pop(idx)
                st.rerun()

# ================================================================================
# PREPARE EXPORT DATA
# ================================================================================
# Define export_flights early so it can be used in Risk & Quality Control section
export_flights = st.session_state.flights if st.session_state.flights else [
    {
        "flight_num": 1,
        "product": product,
        "publisher": publisher,
        "format": ad_format,
        "device": device,
        "start_date": start_date,
        "end_date": end_date,
        "budget": budget,
        "cpm": cpm_rate,
        "volume": impressions,
        "currency": currency_sym,
        "total_cost": total_cost
    }
]

# ================================================================================
# RISK & QUALITY CONTROL
# ================================================================================
st.markdown('<div style="margin-bottom: 1.5rem;"></div>', unsafe_allow_html=True)
st.markdown("## Risk & Quality Control")

product_config = get_product_config(product, publisher, ad_format, device)
freq_cap_text = product_config.get("frequency_cap", "6 imps/day")
delivery_pressure, pressure_explanation = calculate_delivery_pressure(
    st.session_state.flights if st.session_state.flights else export_flights,
    product,
    freq_cap_text
)

# Calculate QC metrics
total_imp = sum(f.get("volume", 0) for f in (st.session_state.flights if st.session_state.flights else export_flights))
first_flight = (st.session_state.flights if st.session_state.flights else export_flights)[0] if (st.session_state.flights or export_flights) else {}
start_dt = first_flight.get("start_date")
end_dt = first_flight.get("end_date")
if isinstance(start_dt, str):
    from datetime import datetime
    start_dt = datetime.strptime(start_dt, "%Y-%m-%d").date()
if isinstance(end_dt, str):
    from datetime import datetime
    end_dt = datetime.strptime(end_dt, "%Y-%m-%d").date()
flight_duration = (end_dt - start_dt).days + 1 if start_dt and end_dt else 0

# ── PER-PRODUCT FORECASTING BREAKDOWN ──────────────────────────────────────────
all_flights_for_forecast = st.session_state.flights if st.session_state.flights else export_flights
if all_flights_for_forecast:
    forecast_groups = {}
    for fl in all_flights_for_forecast:
        fkey = (
            fl.get("product", product),
            fl.get("publisher", publisher),
            fl.get("format", ad_format),
        )
        if fkey not in forecast_groups:
            forecast_groups[fkey] = {"budget": 0, "impressions": 0, "cpm": fl.get("cpm", cpm_rate), "days": 0, "flights": 0}
        forecast_groups[fkey]["budget"] += fl.get("budget", 0)
        forecast_groups[fkey]["impressions"] += fl.get("volume", 0)
        forecast_groups[fkey]["flights"] += 1
        sd, ed = fl.get("start_date"), fl.get("end_date")
        if hasattr(sd, "__sub__") and hasattr(ed, "__sub__"):
            forecast_groups[fkey]["days"] += (ed - sd).days + 1

    forecast_rows = []
    grand_budget = grand_imps = 0
    for (fprod, fpub, ffmt), fdata in forecast_groups.items():
        _fcap = get_freq_cap_numeric(fprod)
        _cpd  = is_cpd_product(fprod)
        if _cpd:
            # CPD row: budget-based, no impressions
            _days = max(1, fdata["days"])
            _daily_rate = fdata["budget"] / _days
            forecast_rows.append({
                "Product": fprod,
                "Publisher": fpub,
                "Format": ffmt,
                "Pricing": "Fixed Fee (CPD)",
                "Budget / Total Fee": f"{fdata['budget']:,.2f} {currency_sym}",
                "Duration": f"{_days} day(s)",
                "Daily Rate": f"{_daily_rate:,.2f} {currency_sym}/day",
                "Total Imps": "SOV — N/A",
                "Est. Daily Avg": "—",
                "Avg Daily Uniques": "N/A (CPD)",
                "Avg Frequency": "N/A (CPD)",
                "Lifetime Uniques": "N/A (CPD)",
                "Flights": fdata["flights"],
            })
        else:
            daily_avg = fdata["impressions"] / max(1, fdata["days"])
            _du = calc_avg_daily_uniques(fdata["impressions"], max(1, fdata["days"]), _fcap)
            _fr = calc_avg_frequency_daily(fdata["impressions"], _du, max(1, fdata["days"]))
            _lu = calc_lifetime_uniques(_du, _turnover_rate, max(1, fdata["days"])) if _du > 0 else 0
            forecast_rows.append({
                "Product": fprod,
                "Publisher": fpub,
                "Format": ffmt,
                "Pricing": f"CPM {fdata['cpm']:,.2f} {currency_sym}",
                "Budget / Total Fee": f"{fdata['budget']:,.2f} {currency_sym}",
                "Duration": f"{max(1,fdata['days'])} day(s)",
                "Daily Rate": "—",
                "Total Imps": f"{fdata['impressions']:,.0f}",
                "Est. Daily Avg": f"{daily_avg:,.0f}",
                "Avg Daily Uniques": f"{_du:,.0f}" if _fcap > 0 else "Uncapped",
                "Avg Frequency": f"{_fr:.2f}" if _fcap > 0 else "Uncapped",
                "Lifetime Uniques": f"{_lu:,.0f}" if _lu > 0 else "—",
                "Flights": fdata["flights"],
            })
        grand_budget += fdata["budget"]
        grand_imps += fdata["impressions"]

    if len(forecast_groups) > 1:
        # Add summary row for multi-product plans
        forecast_rows.append({
            "Product": "📄 TOTAL", "Publisher": "", "Format": "",
            "Pricing": "—",
            "Budget / Total Fee": f"{grand_budget:,.2f} {currency_sym}",
            "Duration": "—",
            "Daily Rate": "—",
            "Total Imps": f"{grand_imps:,.0f}" if grand_imps > 0 else "—",
            "Est. Daily Avg": "—",
            "Avg Daily Uniques": "—",
            "Avg Frequency": "—",
            "Lifetime Uniques": "—",
            "Flights": len(all_flights_for_forecast),
        })
    if forecast_rows:
        st.dataframe(
            pd.DataFrame(forecast_rows), use_container_width=True, hide_index=True,
            column_config={
                "Product": st.column_config.TextColumn(width="large"),
                "Budget": st.column_config.TextColumn(width="medium"),
                "Total Imps": st.column_config.TextColumn(width="medium"),
            }
        )

# ── UNIFIED QC CHECKLIST ────────────────────────────────────────────────────────
_inv_type      = resolve_inventory_type(ad_format)
_inv_tgt_mode  = resolve_inventory_targeting_mode(publisher, ad_format)
_fmt_valid, _fmt_err = validate_deal_format(product, ad_format)
_allowed_sizes = get_allowed_banner_sizes(publisher, device) if ad_format == "Banner" else []

unified_checklist = [
    ("Product",            f"{product} | {publisher} | {ad_format}"),
    ("Priority",           product_config.get("priority", "N/A")),
    ("Line Item Type",     product_config.get("line_item_type", "N/A")),
    ("Revenue Type",       product_config.get("revenue_type", "CPM")),
    ("Frequency Cap",      product_config.get("frequency_cap", "N/A")),
    ("Pacing",             product_config.get("pacing", "Even")),
    # Spec §4.1–4.2: resolved inventory + targeting mode
    ("Inventory Type",     _inv_type),
    ("Inventory Targeting Mode", _inv_tgt_mode),
    ("Inventory Targeting", product_config.get("inventory_targeting",
                            product_config.get("publisher_targeting", "N/A"))),
    ("Supply",             product_config.get("supply", product_config.get("supply_targeting", "N/A"))),
    ("Ad Sizes",           product_config.get("ad_sizes", "N/A")),
    # Spec §4.4: canonical allowed sizes
    ("Allowed Banner Sizes", ", ".join(_allowed_sizes) if _allowed_sizes else "N/A (non-Banner)"),
    # Spec §4.3: format validation
    ("Format Validation",  "✅ Valid" if _fmt_valid else f"❌ INVALID — {_fmt_err}"),
    ("Device Targeting",   product_config.get("device_targeting", "N/A")),
    ("Creative Specs",     product_config.get("creative_specs", "N/A")),
    ("Allow RTB",          str(product_config.get("allow_rtb", "N/A"))),
    ("Underspend Catchup", product_config.get("underspend_catchup", "N/A")),
    ("Geo Targeting",      product_config.get("geo_targeting", "Country Targeting")),
    ("Dates",              f"{start_dt.strftime('%m/%d/%Y') if start_dt else 'N/A'} \u2192 {end_dt.strftime('%m/%d/%Y') if end_dt else 'N/A'}"),
    ("Flight Duration",    f"{flight_duration} days"),
    ("Delivery Risk",      delivery_pressure),
    ("Total Flights",      len(st.session_state.flights if st.session_state.flights else export_flights)),
    # CPD vs CPM pricing rows
    *([("Total Fixed Fee",   f"{budget:,.2f} {currency_sym}"),
       ("Daily Rate",        f"{_cpd_rate:,.2f} {currency_sym}/day"),
       ("Takeover Days",     f"{_flight_days_pricing}"),
       ("Impressions",       "SOV — N/A (Fixed Fee)")] if _takeover else
      [("Total Impressions", f"{total_imp:,.0f}"),
       ("Est. Daily Imps",   f"{total_imp/flight_duration:,.0f}" if flight_duration > 0 else "0"),
       ("CPM",               f"{cpm_rate:.2f} {currency_sym}")]),
    ("Request Date",       datetime.now().strftime("%m/%d/%Y")),
]

# Spec §6 + §8: Forecasting metrics — guarded by CPD check
if not _is_cpd and _freq_cap_num > 0:
    unified_checklist += [
        ("── Audience Forecast ──", ""),
        ("Est. Avg Daily Uniques",  f"{_avg_daily_uniques_est:,.0f}"),
        ("Est. Avg Daily Frequency", f"{_avg_freq_est:.2f}"),
        ("Daily Turnover Rate",     f"{int(_turnover_rate * 100)}%"),
        ("Est. Lifetime Uniques",   f"{_lifetime_uniques_est:,.0f}"),
    ]
elif _is_cpd:
    unified_checklist.append(("Audience Forecast", "❌ N/A — Fixed Fee (CPD) product"))
else:
    unified_checklist.append(("Audience Forecast", "❌ N/A — Uncapped product"))

# Add video-specific rows when applicable
if is_video_product(product, ad_format):
    if video_duration:
        unified_checklist.append(("Creative Duration", video_duration))
    unified_checklist.append(("Target VCR (%)", f"{vcr_target}%"))
    unified_checklist.append(("Est. Completed Views", f"{completed_views:,.0f}" if completed_views > 0 else "—"))
    if completed_views > 0 and budget > 0:
        unified_checklist.append(("Est. Cost / Completed View", f"{currency_sym}{budget / completed_views:.4f}"))

# Display unified table
table_data = []
for item_name, item_value in unified_checklist:
    table_data.append([item_name, str(item_value)])

df = pd.DataFrame(table_data, columns=["Item", "Value"])
st.dataframe(
    df, 
    use_container_width=True, 
    hide_index=True,
    column_config={
        "Item": st.column_config.TextColumn(width="medium"),
        "Value": st.column_config.TextColumn(width="large"),
    }
)
st.markdown('<div style="margin-bottom: 1.5rem;"></div>', unsafe_allow_html=True)

# ================================================================================
# EXPORT MEDIA PLAN
# ================================================================================
st.markdown("## Export Media Plan")
st.markdown('<div style="margin-bottom: 1.5rem;"></div>', unsafe_allow_html=True)

st.write("Download your media plan for client proposal in your preferred format:")
st.markdown('<div style="margin-bottom: 1rem;"></div>', unsafe_allow_html=True)

col_export1, col_export2, col_export3 = st.columns(3)

# IO Excel Export
with col_export1:
    excel_file = create_excel_export(
        export_flights, 
        st.session_state.campaign_details,
        delivery_pressure,
        product_config,
        cpm_rate,
        currency_sym,
        vcr_target=vcr_target if is_video_product(product, ad_format) else None,
        completed_views=completed_views if is_video_product(product, ad_format) else None
    )
    excel_bytes = BytesIO()
    excel_file.save(excel_bytes)
    excel_bytes.seek(0)
    
    _adv_slug   = re.sub(r"[^\w]", "_", advertiser or "Plan")[:20].strip("_")
    _mkt_slug   = re.sub(r"[^\w]", "_", market or "MKT")[:6].strip("_")
    _prod_slug  = re.sub(r"[^\w]", "_", product or "Product")[:14].strip("_")
    _date_slug  = datetime.now().strftime("%Y%m%d")
    _base_name  = f"{_adv_slug}_{_mkt_slug}_{_prod_slug}_{_date_slug}"

    st.download_button(
        label="📊 IO (Excel)",
        data=excel_bytes.getvalue(),
        file_name=f"{_base_name}_IO.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Microsoft-style Excel file with campaign details and forecasting"
    )

# IO PDF Export
with col_export2:
    pdf_buffer = create_pdf_export(
        export_flights, 
        st.session_state.campaign_details,
        delivery_pressure,
        product_config,
        cpm_rate,
        currency_sym,
        vcr_target=vcr_target if is_video_product(product, ad_format) else None,
        completed_views=completed_views if is_video_product(product, ad_format) else None
    )
    
    st.download_button(
        label="📄 IO (PDF)",
        data=pdf_buffer.getvalue(),
        file_name=f"{_base_name}_IO.pdf",
        mime="application/pdf",
        use_container_width=True,
        help="PDF version of the media plan"
    )

# Internal-Facing Excel Export
with col_export3:
    # Create internal-facing Excel with targeting setup
    internal_file = create_internal_excel_export(
        export_flights, st.session_state.campaign_details, delivery_pressure,
        product_config, unified_checklist
    )
    internal_bytes = BytesIO()
    internal_file.save(internal_bytes)
    internal_bytes.seek(0)
    
    st.download_button(
        label="🔒 Internal Setup (Excel)",
        data=internal_bytes.getvalue(),
        file_name=f"{_base_name}_InternalSetup.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Internal-facing Excel with targeting setup and QC checklist"
    )

st.divider()

st.markdown(f"""
<p style='text-align: center; font-size: 0.85rem; color: #666;'>
Monetize Guaranteed Media Planner • Seat 280 • Generated {datetime.now().strftime('%b %d, %Y at %I:%M %p')}
</p>
""", unsafe_allow_html=True)
